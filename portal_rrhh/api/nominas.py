# -*- coding: utf-8 -*-
# Copyright (c) 2026, Grupo ATU and contributors
# For license information, please see license.txt

import frappe
from frappe import _
from frappe.utils import flt, today, getdate
import json

# Importar funciones compartidas del DocType
from hrms_integrations.hrms_integrations.doctype.liquidacion_nomina.liquidacion_nomina import (
	calcular_horas_mes_desde_calendario,
	get_job_offers_docentes,
	get_modificaciones_docentes,
	get_courses_from_job_offer,
	get_courses_from_modificacion,
	existe_liquidacion,
	get_ultima_liquidacion_empleado,
	MONTH_MAP
)

DOCTYPE_NAME = "Liquidacion Nomina"


@frappe.whitelist()
def get_prevision_mes(mes, año, employee=None, course=None):
	"""
	Obtiene la previsión de liquidaciones para un mes específico.
	
	Busca Job Offers y Modificaciones RRHH que:
	1. Estén actualmente en Alta (workflow_state = 'Alta')
	2. O que hayan estado activos en el periodo desde la última liquidación
	
	Excluye los que ya tienen liquidación creada para el mes/año.
	"""
	año = int(año)
	previsiones = []
	documentos_procesados = set()  # Para evitar duplicados
	
	# Obtener Job Offers (activos y con histórico pendiente)
	job_offers = get_job_offers_docentes(
		employee=employee, 
		mes=mes, 
		año=año, 
		incluir_historico=True
	)
	
	for jo in job_offers:
		# Obtener courses del Job Offer
		courses = get_courses_from_job_offer(jo.name)
		
		for course_name in courses:
			# Evitar duplicados
			doc_key = f"JO-{jo.name}-{course_name}"
			if doc_key in documentos_procesados:
				continue
			documentos_procesados.add(doc_key)
			
			# Filtrar por course si se especificó (búsqueda parcial)
			if course and course.lower() not in course_name.lower():
				continue
			
			# Verificar si ya existe liquidación para este empleado + curso + mes
			if existe_liquidacion(jo.name, course_name, mes, año, employee=jo.employee):
				continue
			
			# Calcular previsión
			prevision = calcular_prevision_liquidacion(
				employee=jo.employee,
				employee_name=jo.employee_name,
				dni_nie=jo.dni_nie,
				designation=jo.designation,
				company=jo.company,
				source_type="Job Offer",
				source_doc=jo.name,
				course=course_name,
				mes=mes,
				año=año,
				precio_hora=jo.precio_hora,
				workflow_state=jo.workflow_state,
				fecha_inicio=jo.fecha_inicio,
				fecha_fin=jo.fecha_fin
			)
			
			if prevision:
				previsiones.append(prevision)
	
	# Obtener Modificaciones RRHH (activas y con histórico pendiente)
	modificaciones = get_modificaciones_docentes(
		employee=employee, 
		mes=mes, 
		año=año, 
		incluir_historico=True
	)
	
	for mod in modificaciones:
		# Obtener courses de la Modificación
		courses = get_courses_from_modificacion(mod.name)
		
		for course_name in courses:
			# Evitar duplicados
			doc_key = f"MOD-{mod.name}-{course_name}"
			if doc_key in documentos_procesados:
				continue
			documentos_procesados.add(doc_key)
			
			# Filtrar por course si se especificó (búsqueda parcial)
			if course and course.lower() not in course_name.lower():
				continue
			
			# Verificar si ya existe liquidación para este empleado + curso + mes
			if existe_liquidacion(mod.name, course_name, mes, año, employee=mod.employee):
				continue
			
			# Calcular previsión
			prevision = calcular_prevision_liquidacion(
				employee=mod.employee,
				employee_name=mod.employee_name,
				dni_nie=mod.dni_nie,
				designation=mod.designation,
				company=mod.company,
				source_type="Modificaciones RRHH",
				source_doc=mod.name,
				course=course_name,
				mes=mes,
				año=año,
				precio_hora=mod.precio_hora,
				workflow_state=mod.workflow_state,
				fecha_inicio=mod.fecha_inicio,
				fecha_fin=mod.fecha_fin
			)
			
			if prevision:
				previsiones.append(prevision)
	
	# Calcular resumen
	resumen = calcular_resumen(previsiones)
	resumen['total_documentos'] = len(documentos_procesados)
	
	return {
		'previsiones': previsiones,
		'resumen': resumen
	}


@frappe.whitelist()
def crear_liquidaciones_mes(mes, año, liquidaciones_data):
	"""
	Crea las liquidaciones para el mes seleccionado en estado BORRADOR.
	Las liquidaciones se crean sin submit para permitir ajustes manuales.
	Luego se validan con validar_liquidaciones().
	"""
	if isinstance(liquidaciones_data, str):
		liquidaciones_data = json.loads(liquidaciones_data)
	
	año = int(año)
	creadas = 0
	errores = []
	liquidaciones_creadas = []
	
	for liq_data in liquidaciones_data:
		try:
			# Verificar si ya existe
			existing = frappe.db.exists(DOCTYPE_NAME, {
				"employee": liq_data['employee'],
				"course": liq_data['course'],
				"mes": mes,
				"año": año,
				"docstatus": ["!=", 2]
			})
			
			if existing:
				errores.append(f"{liq_data['employee_name']}: Ya existe liquidación")
				continue
			
			# Crear liquidación en estado BORRADOR (sin submit)
			doc = frappe.new_doc(DOCTYPE_NAME)
			doc.employee = liq_data['employee']
			doc.mes = mes
			doc.año = año
			doc.company = liq_data['company']
			doc.source_document_type = liq_data['source_type']
			doc.source_document = liq_data['source_doc']
			doc.course = liq_data['course']
			doc.precio_hora = liq_data['precio_hora']
			doc.horas_extras = flt(liq_data.get('horas_extras', 0))
			doc.precio_hora_extra = flt(liq_data.get('precio_hora_extra', 0))
			doc.es_ultimo_mes = liq_data.get('es_ultimo_mes', 0)
			doc.estado = "Borrador"  # Estado Borrador para permitir edición
			
			doc.insert()
			# NO hacemos submit() - queda en docstatus=0 (Borrador)
			creadas += 1
			liquidaciones_creadas.append(doc.name)
			
		except Exception as e:
			errores.append(f"{liq_data.get('employee_name', 'Unknown')}: {str(e)}")
			frappe.log_error(f"Error creando liquidación: {str(e)}", "Crear Liquidación")
	
	return {
		'success': True,
		'creadas': creadas,
		'errores': errores,
		'liquidaciones': liquidaciones_creadas,
		'message': _("Se crearon {0} liquidaciones en borrador. Revísalas y valídalas.").format(creadas)
	}


@frappe.whitelist()
def generar_liquidaciones_masivas(mes, año):
	"""
	Genera todas las liquidaciones pendientes para el mes especificado.
	
	Esta función:
	1. Obtiene todas las Job Offers y Modificaciones RRHH pendientes
	2. Filtra las que ya tienen liquidación
	3. Crea las liquidaciones en lote
	4. Retorna un resumen del proceso
	
	Args:
		mes: Mes para generar liquidaciones (ej: 'Febrero')
		año: Año para generar liquidaciones
	
	Returns:
		dict con resumen del proceso (creadas, omitidas, errores, por_docente)
	"""
	año = int(año)
	
	# Obtener todas las previsiones pendientes
	resultado_prevision = get_prevision_mes(mes, año)
	previsiones = resultado_prevision.get('previsiones', [])
	
	if not previsiones:
		return {
			'success': True,
			'creadas': 0,
			'omitidas': 0,
			'errores': [],
			'por_docente': {},
			'message': _("No hay liquidaciones pendientes para {0} {1}").format(mes, año)
		}
	
	# Agrupar por docente para el reporte
	por_docente = {}
	creadas = 0
	omitidas = 0
	errores = []
	
	for prev in previsiones:
		employee_name = prev.get('employee_name', 'Desconocido')
		employee = prev.get('employee')
		
		if employee_name not in por_docente:
			por_docente[employee_name] = {
				'employee': employee,
				'liquidaciones': [],
				'creadas': 0,
				'errores': 0
			}
		
		try:
			# Verificar nuevamente si ya existe (por si se creó entre la previsión y ahora)
			if existe_liquidacion(prev['source_doc'], prev['course'], mes, año, employee=prev['employee']):
				omitidas += 1
				continue
			
			# Crear liquidación
			doc = frappe.new_doc(DOCTYPE_NAME)
			doc.employee = prev['employee']
			doc.mes = mes
			doc.año = año
			doc.company = prev['company']
			doc.source_document_type = prev['source_type']
			doc.source_document = prev['source_doc']
			doc.course = prev['course']
			doc.precio_hora = prev['precio_hora']
			doc.horas_extras = flt(prev.get('horas_extras', 0))
			doc.precio_hora_extra = flt(prev.get('precio_hora_extra', 0))
			doc.es_ultimo_mes = prev.get('es_ultimo_mes', 0)
			doc.estado = "Borrador"  # Estado Borrador para permitir edición
			
			doc.insert()
			# NO hacemos submit() - queda en docstatus=0 (Borrador)
			
			creadas += 1
			por_docente[employee_name]['creadas'] += 1
			por_docente[employee_name]['liquidaciones'].append({
				'name': doc.name,
				'course': prev['course'],
				'total': doc.total  # Usar el total calculado por el DocType
			})
			
		except frappe.exceptions.ValidationError as e:
			error_msg = f"{employee_name} - {prev.get('course', 'N/A')}: {str(e)}"
			errores.append(error_msg)
			por_docente[employee_name]['errores'] += 1
		except Exception as e:
			error_msg = f"{employee_name} - {prev.get('course', 'N/A')}: {str(e)}"
			errores.append(error_msg)
			por_docente[employee_name]['errores'] += 1
			frappe.log_error(message=str(e)[:500], title="Liquidaciones Masivas Error")
	
	# Resumen por docente (solo los que tuvieron liquidaciones)
	resumen_docentes = {
		nombre: datos for nombre, datos in por_docente.items() 
		if datos['creadas'] > 0 or datos['errores'] > 0
	}
	
	return {
		'success': True,
		'creadas': creadas,
		'omitidas': omitidas,
		'errores': errores,
		'total_docentes': len(resumen_docentes),
		'por_docente': resumen_docentes,
		'message': _("Se crearon {0} liquidaciones en borrador para {1} docentes. Revísalas y valídalas.").format(creadas, len(resumen_docentes))
	}


@frappe.whitelist()
def get_liquidaciones_mes(mes, año, employee=None, course=None, estado=None):
	"""
	Obtiene las liquidaciones ya creadas para un mes.
	Incluye tanto borradores (docstatus=0) como validadas (docstatus=1).
	"""
	filters = {
		"mes": mes,
		"año": int(año),
		"docstatus": ["!=", 2]  # Excluir canceladas
	}
	
	if estado:
		filters["estado"] = estado
	
	liquidaciones = frappe.get_all(DOCTYPE_NAME,
		filters=filters,
		fields=[
			"name", "employee", "employee_name", "dni_nie", "designation",
			"course", "company", "mes", "año",
			"horas_normales", "horas_extras", "total_horas", "dias_trabajados",
			"precio_hora", "precio_hora_extra",
			"bruto", "vacaciones_mes", "bruto_menos_vacaciones",
			"vacaciones_acumuladas", "base_ss", "importe_ss", "total",
			"estado", "fecha_liquidacion", "fecha_envio_asesoria", "fecha_pago",
			"es_ultimo_mes", "docstatus",
			"source_document", "source_document_type"
		],
		order_by="docstatus asc, employee_name asc, course asc"
	)
	
	# Añadir course_display a cada liquidación
	for liq in liquidaciones:
		liq['course_display'] = frappe.db.get_value('Course', liq['course'], 'custom_display_identifier') or liq['course']
	
	# Calcular resumen
	resumen = calcular_resumen(liquidaciones)
	
	return {
		'liquidaciones': liquidaciones,
		'resumen': resumen
	}


@frappe.whitelist()
def actualizar_horas_extras(liquidacion_id, horas_extras, precio_hora_extra=None):
	"""
	Actualiza las horas extras de una liquidación.
	Solo se puede hacer si no está submitted.
	"""
	doc = frappe.get_doc(DOCTYPE_NAME, liquidacion_id)
	
	if doc.docstatus == 1:
		frappe.throw(_("No se pueden modificar liquidaciones ya enviadas"))
	
	doc.horas_extras = flt(horas_extras)
	if precio_hora_extra:
		doc.precio_hora_extra = flt(precio_hora_extra)
	
	doc.save()
	
	return {
		'success': True,
		'message': _("Horas extras actualizadas"),
		'liquidacion': doc.as_dict()
	}


@frappe.whitelist()
def editar_liquidacion(liquidacion_id, campos):
	"""
	Edita una liquidación en estado Borrador y recalcula automáticamente.
	
	Args:
		liquidacion_id: ID de la liquidación
		campos: dict con los campos a actualizar:
			- horas_normales
			- horas_extras
			- precio_hora
			- precio_hora_extra
			- es_ultimo_mes
	
	Returns:
		dict con la liquidación actualizada
	"""
	if isinstance(campos, str):
		campos = json.loads(campos)
	
	doc = frappe.get_doc(DOCTYPE_NAME, liquidacion_id)
	
	if doc.docstatus == 1:
		frappe.throw(_("No se pueden modificar liquidaciones ya validadas"))
	
	if doc.docstatus == 2:
		frappe.throw(_("No se pueden modificar liquidaciones canceladas"))
	
	# Actualizar campos editables
	campos_editables = ['horas_normales', 'horas_extras', 'precio_hora', 'precio_hora_extra', 'es_ultimo_mes', 'vacaciones_acumuladas']
	
	for campo in campos_editables:
		if campo in campos and campos[campo] is not None:
			if campo == 'es_ultimo_mes':
				setattr(doc, campo, 1 if campos[campo] else 0)
			elif campo == 'vacaciones_acumuladas':
				# Solo permitir vacaciones acumuladas si es último mes
				if campos.get('es_ultimo_mes') or doc.es_ultimo_mes:
					setattr(doc, campo, flt(campos[campo]))
			else:
				setattr(doc, campo, flt(campos[campo]))
	
	# Guardar - el método validate() del DocType recalculará automáticamente
	doc.save()
	
	return {
		'success': True,
		'message': _("Liquidación actualizada y recalculada"),
		'liquidacion': doc.as_dict()
	}


@frappe.whitelist()
def validar_liquidaciones(liquidaciones_ids):
	"""
	Valida (submit) las liquidaciones seleccionadas.
	Cambia el estado de Borrador a Liquidado.
	
	Args:
		liquidaciones_ids: lista de IDs de liquidaciones a validar
	
	Returns:
		dict con resumen de validadas y errores
	"""
	if isinstance(liquidaciones_ids, str):
		liquidaciones_ids = json.loads(liquidaciones_ids)
	
	validadas = 0
	errores = []
	
	for liq_id in liquidaciones_ids:
		try:
			doc = frappe.get_doc(DOCTYPE_NAME, liq_id)
			
			if doc.docstatus == 1:
				errores.append(f"{doc.employee_name}: Ya está validada")
				continue
			
			if doc.docstatus == 2:
				errores.append(f"{doc.employee_name}: Está cancelada")
				continue
			
			# Actualizar estado y hacer submit
			doc.estado = "Liquidado"
			doc.save()
			doc.submit()
			validadas += 1
			
		except Exception as e:
			errores.append(f"{liq_id}: {str(e)}")
			frappe.log_error(f"Error validando liquidación {liq_id}: {str(e)}", "Validar Liquidaciones")
	
	return {
		'success': True,
		'validadas': validadas,
		'errores': errores,
		'message': _("Se validaron {0} liquidaciones").format(validadas)
	}


@frappe.whitelist()
def obtener_vacaciones_acumuladas(employee, course, año):
	"""
	Obtiene las vacaciones acumuladas de un empleado para un curso/año.
	Solo cuenta las liquidaciones VALIDADAS (docstatus=1).
	
	Args:
		employee: ID del empleado
		course: Nombre del curso
		año: Año
	
	Returns:
		dict con total de vacaciones acumuladas y detalle por mes
	"""
	año = int(año)
	
	liquidaciones = frappe.db.sql("""
		SELECT name, mes, vacaciones_mes, docstatus, estado
		FROM `tabLiquidacion Nomina`
		WHERE employee = %(employee)s
		AND course = %(course)s
		AND año = %(año)s
		AND docstatus = 1
		ORDER BY 
			CASE mes
				WHEN 'Enero' THEN 1
				WHEN 'Febrero' THEN 2
				WHEN 'Marzo' THEN 3
				WHEN 'Abril' THEN 4
				WHEN 'Mayo' THEN 5
				WHEN 'Junio' THEN 6
				WHEN 'Julio' THEN 7
				WHEN 'Agosto' THEN 8
				WHEN 'Septiembre' THEN 9
				WHEN 'Octubre' THEN 10
				WHEN 'Noviembre' THEN 11
				WHEN 'Diciembre' THEN 12
			END
	""", {
		"employee": employee,
		"course": course,
		"año": año
	}, as_dict=True)
	
	total_acumulado = sum(flt(liq.vacaciones_mes) for liq in liquidaciones)
	
	return {
		'success': True,
		'total_acumulado': round(total_acumulado, 2),
		'meses_liquidados': len(liquidaciones),
		'detalle': liquidaciones
	}


@frappe.whitelist()
def marcar_como_enviado(liquidaciones_ids):
	"""
	Marca las liquidaciones seleccionadas como enviadas a asesoría.
	"""
	if isinstance(liquidaciones_ids, str):
		liquidaciones_ids = json.loads(liquidaciones_ids)
	
	fecha_envio = today()
	actualizadas = 0
	errores = []
	
	for liq_id in liquidaciones_ids:
		try:
			doc = frappe.get_doc(DOCTYPE_NAME, liq_id)
			
			if doc.docstatus != 1:
				errores.append(f"{liq_id}: No está liquidada")
				continue
			
			doc.estado = "Enviado a Asesoría"
			doc.fecha_envio_asesoria = fecha_envio
			doc.save()
			actualizadas += 1
			
		except Exception as e:
			errores.append(f"{liq_id}: {str(e)}")
			frappe.log_error(f"Error marcando como enviado {liq_id}: {str(e)}")
	
	return {
		'success': True,
		'actualizadas': actualizadas,
		'errores': errores,
		'message': _("Se marcaron {0} liquidaciones como enviadas").format(actualizadas)
	}


@frappe.whitelist()
def marcar_como_pagado(liquidaciones_ids, fecha_pago=None):
	"""
	Marca las liquidaciones como pagadas.
	"""
	if isinstance(liquidaciones_ids, str):
		liquidaciones_ids = json.loads(liquidaciones_ids)
	
	if not fecha_pago:
		fecha_pago = today()
	
	actualizadas = 0
	errores = []
	
	for liq_id in liquidaciones_ids:
		try:
			doc = frappe.get_doc(DOCTYPE_NAME, liq_id)
			
			if doc.docstatus != 1:
				errores.append(f"{liq_id}: No está liquidada")
				continue
			
			doc.estado = "Pagado"
			doc.fecha_pago = fecha_pago
			doc.save()
			actualizadas += 1
			
		except Exception as e:
			errores.append(f"{liq_id}: {str(e)}")
			frappe.log_error(f"Error marcando como pagado {liq_id}: {str(e)}")
	
	return {
		'success': True,
		'actualizadas': actualizadas,
		'errores': errores,
		'message': _("Se marcaron {0} liquidaciones como pagadas").format(actualizadas)
	}


@frappe.whitelist()
def generar_reporte_excel(mes, año, liquidaciones_ids=None):
	"""
	Genera el reporte Excel para enviar a la asesoría.
	Devuelve la URL del archivo generado.
	"""
	if isinstance(liquidaciones_ids, str):
		liquidaciones_ids = json.loads(liquidaciones_ids)
	
	filters = {
		"mes": mes,
		"año": int(año),
		"docstatus": 1
	}
	
	if liquidaciones_ids:
		filters["name"] = ["in", liquidaciones_ids]
	
	liquidaciones = frappe.get_all(DOCTYPE_NAME,
		filters=filters,
		fields="*",
		order_by="company asc, employee_name asc"
	)
	
	if not liquidaciones:
		frappe.throw(_("No hay liquidaciones para exportar"))
	
	# Generar Excel
	try:
		from openpyxl import Workbook
		from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
		from openpyxl.utils import get_column_letter
		import io
		
		wb = Workbook()
		ws = wb.active
		ws.title = f"Liquidaciones {mes} {año}"
		
		# Estilos
		header_font = Font(bold=True, color="FFFFFF")
		header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
		header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
		thin_border = Border(
			left=Side(style='thin'),
			right=Side(style='thin'),
			top=Side(style='thin'),
			bottom=Side(style='thin')
		)
		currency_format = '#,##0.00 €'
		
		# Headers
		headers = [
			"Empresa", "Empleado", "DNI/NIE", "Designation", "Course",
			"Horas Normales", "Horas Extras", "Total Horas", "Días Trabajados",
			"Precio/Hora", "Precio/Hora Extra",
			"Bruto", "Vacaciones Mes", "Bruto - Vacaciones",
			"Base SS", "Importe SS", "TOTAL",
			"Estado", "Fecha Liquidación"
		]
		
		for col_num, header in enumerate(headers, 1):
			cell = ws.cell(row=1, column=col_num, value=header)
			cell.font = header_font
			cell.fill = header_fill
			cell.alignment = header_alignment
			cell.border = thin_border
		
		# Datos
		for row_num, liq in enumerate(liquidaciones, 2):
			row_data = [
				liq.company,
				liq.employee_name,
				liq.dni_nie,
				liq.designation,
				liq.course,
				liq.horas_normales,
				liq.horas_extras,
				liq.total_horas,
				liq.dias_trabajados,
				liq.precio_hora,
				liq.precio_hora_extra,
				liq.bruto,
				liq.vacaciones_mes,
				liq.bruto_menos_vacaciones,
				liq.base_ss,
				liq.importe_ss,
				liq.total,
				liq.estado,
				str(liq.fecha_liquidacion) if liq.fecha_liquidacion else ""
			]
			
			for col_num, value in enumerate(row_data, 1):
				cell = ws.cell(row=row_num, column=col_num, value=value)
				cell.border = thin_border
				
				# Formato de moneda para columnas de importes
				if col_num in [10, 11, 12, 13, 14, 15, 16, 17]:
					cell.number_format = currency_format
		
		# Fila de totales
		total_row = len(liquidaciones) + 2
		ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True)
		
		# Sumar columnas numéricas
		sum_columns = {
			6: sum(flt(l.horas_normales) for l in liquidaciones),
			7: sum(flt(l.horas_extras) for l in liquidaciones),
			8: sum(flt(l.total_horas) for l in liquidaciones),
			12: sum(flt(l.bruto) for l in liquidaciones),
			13: sum(flt(l.vacaciones_mes) for l in liquidaciones),
			14: sum(flt(l.bruto_menos_vacaciones) for l in liquidaciones),
			15: sum(flt(l.base_ss) for l in liquidaciones),
			16: sum(flt(l.importe_ss) for l in liquidaciones),
			17: sum(flt(l.total) for l in liquidaciones),
		}
		
		for col, value in sum_columns.items():
			cell = ws.cell(row=total_row, column=col, value=value)
			cell.font = Font(bold=True)
			cell.border = thin_border
			if col >= 10:
				cell.number_format = currency_format
		
		# Ajustar anchos de columna
		column_widths = [15, 25, 12, 15, 30, 12, 12, 12, 12, 12, 12, 12, 12, 15, 12, 12, 12, 15, 15]
		for i, width in enumerate(column_widths, 1):
			ws.column_dimensions[get_column_letter(i)].width = width
		
		# Guardar en buffer
		output = io.BytesIO()
		wb.save(output)
		output.seek(0)
		
		# Guardar como archivo en Frappe
		filename = f"Liquidaciones_{mes}_{año}.xlsx"
		file_doc = frappe.get_doc({
			"doctype": "File",
			"file_name": filename,
			"content": output.getvalue(),
			"is_private": 1
		})
		file_doc.save(ignore_permissions=True)
		
		return {
			'success': True,
			'file_url': file_doc.file_url,
			'filename': filename,
			'message': _("Reporte generado correctamente")
		}
		
	except ImportError:
		frappe.throw(_("El módulo openpyxl no está instalado. Ejecute: pip install openpyxl"))
	except Exception as e:
		frappe.log_error(f"Error generando Excel: {str(e)}", "Generar Reporte Excel")
		frappe.throw(_("Error generando el reporte: {0}").format(str(e)))


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def calcular_prevision_liquidacion(employee, employee_name, dni_nie, designation, company,
									source_type, source_doc, course, mes, año, precio_hora,
									workflow_state=None, fecha_inicio=None, fecha_fin=None):
	"""
	Calcula la previsión de liquidación para un empleado/course/mes.
	
	Args:
		employee: ID del empleado
		employee_name: Nombre del empleado
		dni_nie: DNI/NIE del empleado
		designation: Cargo/Designation
		company: Empresa
		source_type: Tipo de documento origen (Job Offer / Modificaciones RRHH)
		source_doc: ID del documento origen
		course: Nombre del curso
		mes: Mes de la liquidación
		año: Año de la liquidación
		precio_hora: Precio por hora
		workflow_state: Estado actual del workflow (Alta, Baja, etc.)
		fecha_inicio: Fecha de inicio del contrato
		fecha_fin: Fecha de fin del contrato
	"""
	# Obtener custom_display_identifier del curso
	course_display = frappe.db.get_value('Course', course, 'custom_display_identifier') or course
	
	# 1. Calcular horas desde el calendario
	horas_data = calcular_horas_mes_desde_calendario(course, mes, año)
	
	if not horas_data or horas_data['total_horas'] == 0:
		return None
	
	horas_normales = horas_data['total_horas']
	dias_trabajados = horas_data['dias_trabajados']
	
	# 2. Calcular importes
	importe_horas_normales = horas_normales * flt(precio_hora)
	bruto = importe_horas_normales
	
	# 3. Calcular vacaciones (8.33%)
	vacaciones_mes = round(bruto * 0.0833, 2)
	
	# 4. Calcular bruto menos vacaciones
	bruto_menos_vacaciones = bruto - vacaciones_mes
	
	# 5. Calcular SS (32.07%)
	importe_ss = round(bruto_menos_vacaciones * 0.3207, 2)
	
	# 6. Calcular total
	total = bruto_menos_vacaciones + importe_ss
	
	return {
		'employee': employee,
		'employee_name': employee_name,
		'dni_nie': dni_nie,
		'designation': designation,
		'company': company,
		'source_type': source_type,
		'source_doc': source_doc,
		'course': course,
		'course_display': course_display,
		'mes': mes,
		'año': año,
		'horas_normales': horas_normales,
		'horas_extras': 0,
		'total_horas': horas_normales,
		'dias_trabajados': dias_trabajados,
		'precio_hora': precio_hora,
		'precio_hora_extra': round(flt(precio_hora) * 1.2, 2),
		'importe_horas_normales': importe_horas_normales,
		'importe_horas_extras': 0,
		'bruto': bruto,
		'vacaciones_mes': vacaciones_mes,
		'bruto_menos_vacaciones': bruto_menos_vacaciones,
		'vacaciones_acumuladas': 0,
		'base_ss': bruto_menos_vacaciones,
		'importe_ss': importe_ss,
		'total': total,
		'estado': 'Previsión',
		'es_ultimo_mes': 0,
		'workflow_state': workflow_state,
		'fecha_inicio': str(fecha_inicio) if fecha_inicio else None,
		'fecha_fin': str(fecha_fin) if fecha_fin else None
	}


def calcular_resumen(items):
	"""Calcula el resumen de una lista de liquidaciones/previsiones"""
	if not items:
		return {
			'total_docentes': 0,
			'total_liquidaciones': 0,
			'total_horas': 0,
			'total_bruto': 0,
			'total_ss': 0,
			'total_pagar': 0
		}
	
	# Obtener employee de cada item (puede ser dict o frappe._dict)
	employees = set()
	for item in items:
		emp = item.get('employee') if isinstance(item, dict) else item.employee
		if emp:
			employees.add(emp)
	
	return {
		'total_docentes': len(employees),
		'total_liquidaciones': len(items),
		'total_horas': sum(flt(item.get('total_horas') if isinstance(item, dict) else item.total_horas) for item in items),
		'total_bruto': sum(flt(item.get('bruto') if isinstance(item, dict) else item.bruto) for item in items),
		'total_ss': sum(flt(item.get('importe_ss') if isinstance(item, dict) else item.importe_ss) for item in items),
		'total_pagar': sum(flt(item.get('total') if isinstance(item, dict) else item.total) for item in items)
	}


@frappe.whitelist()
def enviar_liquidaciones_asesoria(liquidaciones_ids, mes, año):
	"""
	Envía liquidaciones validadas a los usuarios con rol Asesoría.
	Genera un Excel profesional y lo envía por email.
	Marca las liquidaciones como 'Enviado a Asesoría'.
	"""
	import io
	from openpyxl import Workbook
	from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
	from openpyxl.utils import get_column_letter
	
	if isinstance(liquidaciones_ids, str):
		liquidaciones_ids = json.loads(liquidaciones_ids)
	
	if not liquidaciones_ids:
		frappe.throw(_("No se han seleccionado liquidaciones para enviar"))
	
	# Obtener liquidaciones con datos completos
	liquidaciones_data = []
	for liq_id in liquidaciones_ids:
		doc = frappe.get_doc(DOCTYPE_NAME, liq_id)
		
		# Solo permitir enviar liquidaciones validadas (docstatus=1)
		if doc.docstatus != 1:
			frappe.throw(_("Solo se pueden enviar liquidaciones validadas. '{0}' no está validada.").format(liq_id))
		
		# Obtener provincia y empresa del documento origen
		provincia = ""
		empresa = doc.company or ""
		
		if doc.source_document_type == "Job Offer" and doc.source_document:
			jo_data = frappe.db.get_value("Job Offer", doc.source_document, 
				["custom_provincia", "company"], as_dict=True)
			if jo_data:
				provincia = jo_data.get("custom_provincia") or ""
				empresa = jo_data.get("company") or empresa
		elif doc.source_document_type == "Modificaciones RRHH" and doc.source_document:
			mr_data = frappe.db.get_value("Modificaciones RRHH", doc.source_document,
				["custom_provincia", "company"], as_dict=True)
			if mr_data:
				provincia = mr_data.get("custom_provincia") or ""
				empresa = mr_data.get("company") or empresa
		
		# Obtener course_display
		course_display = frappe.db.get_value('Course', doc.course, 'custom_display_identifier') or doc.course
		
		liquidaciones_data.append({
			'name': doc.name,
			'mes': doc.mes,
			'año': doc.año,
			'employee_name': doc.employee_name,
			'dni_nie': doc.dni_nie,
			'designation': doc.designation,
			'company': empresa,
			'provincia': provincia,
			'course': doc.course,
			'course_display': course_display,
			'horas_normales': doc.horas_normales,
			'horas_extras': doc.horas_extras,
			'total_horas': doc.total_horas,
			'precio_hora': doc.precio_hora,
			'bruto': doc.bruto,
			'vacaciones_mes': doc.vacaciones_mes,
			'bruto_menos_vacaciones': doc.bruto_menos_vacaciones,
			'vacaciones_acumuladas': doc.vacaciones_acumuladas,
			'base_ss': doc.base_ss,
			'importe_ss': doc.importe_ss,
			'total': doc.total,
			'es_ultimo_mes': doc.es_ultimo_mes,
			'source_document_type': doc.source_document_type,
			'source_document': doc.source_document
		})
	
	# Crear Excel
	wb = Workbook()
	ws = wb.active
	ws.title = f"Liquidaciones {mes} {año}"
	
	# Estilos
	header_font = Font(bold=True, color="FFFFFF", size=11)
	header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
	header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
	
	currency_format = '#,##0.00 €'
	number_format = '#,##0.00'
	
	thin_border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)
	
	alt_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
	ultimo_mes_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
	
	# Título
	ws.merge_cells('A1:R1')
	title_cell = ws['A1']
	title_cell.value = f"LIQUIDACIONES DE NÓMINAS - {mes.upper()} {año}"
	title_cell.font = Font(bold=True, size=14, color="1F4E79")
	title_cell.alignment = Alignment(horizontal="center", vertical="center")
	ws.row_dimensions[1].height = 30
	
	# Subtítulo con fecha de generación
	ws.merge_cells('A2:R2')
	subtitle_cell = ws['A2']
	subtitle_cell.value = f"Generado el {frappe.utils.now_datetime().strftime('%d/%m/%Y %H:%M')} - Total: {len(liquidaciones_data)} liquidaciones"
	subtitle_cell.font = Font(italic=True, size=10, color="666666")
	subtitle_cell.alignment = Alignment(horizontal="center")
	ws.row_dimensions[2].height = 20
	
	# Headers
	headers = [
		("Mes", 10),
		("Año", 8),
		("Docente", 25),
		("DNI/NIE", 12),
		("Cargo", 15),
		("Empresa", 20),
		("Provincia", 15),
		("Curso", 30),
		("Horas N.", 10),
		("Horas E.", 10),
		("Total H.", 10),
		("€/Hora", 10),
		("Bruto+V", 12),
		("Vacaciones", 12),
		("Bruto", 12),
		("SS Empresa", 12),
		("TOTAL", 12),
		("Últ. Mes", 10)
	]
	
	header_row = 4
	for col_idx, (header_name, width) in enumerate(headers, 1):
		cell = ws.cell(row=header_row, column=col_idx, value=header_name)
		cell.font = header_font
		cell.fill = header_fill
		cell.alignment = header_alignment
		cell.border = thin_border
		ws.column_dimensions[get_column_letter(col_idx)].width = width
	
	ws.row_dimensions[header_row].height = 25
	
	# Datos
	data_alignment = Alignment(vertical="center")
	currency_alignment = Alignment(horizontal="right", vertical="center")
	
	for row_idx, liq in enumerate(liquidaciones_data, header_row + 1):
		row_data = [
			liq['mes'],
			liq['año'],
			liq['employee_name'],
			liq['dni_nie'],
			liq['designation'],
			liq['company'],
			liq['provincia'],
			liq['course_display'],
			liq['horas_normales'],
			liq['horas_extras'],
			liq['total_horas'],
			liq['precio_hora'],
			liq['bruto'],
			liq['vacaciones_mes'],
			liq['bruto_menos_vacaciones'],
			liq['importe_ss'],
			liq['total'],
			"Sí" if liq['es_ultimo_mes'] else "No"
		]
		
		for col_idx, value in enumerate(row_data, 1):
			cell = ws.cell(row=row_idx, column=col_idx, value=value)
			cell.border = thin_border
			cell.alignment = data_alignment
			
			# Formato numérico/moneda
			if col_idx in [9, 10, 11]:  # Horas
				cell.number_format = number_format
			elif col_idx in [12, 13, 14, 15, 16, 17]:  # Importes
				cell.number_format = currency_format
				cell.alignment = currency_alignment
			
			# Colores alternados o último mes
			if liq['es_ultimo_mes']:
				cell.fill = ultimo_mes_fill
			elif row_idx % 2 == 0:
				cell.fill = alt_fill
	
	# Fila de totales
	total_row = header_row + len(liquidaciones_data) + 1
	ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True)
	ws.cell(row=total_row, column=1).border = thin_border
	
	for col_idx in range(2, 9):
		ws.cell(row=total_row, column=col_idx).border = thin_border
	
	# Sumar columnas numéricas
	sum_cols = {
		9: sum(l['horas_normales'] for l in liquidaciones_data),
		10: sum(l['horas_extras'] for l in liquidaciones_data),
		11: sum(l['total_horas'] for l in liquidaciones_data),
		13: sum(l['bruto'] for l in liquidaciones_data),
		14: sum(l['vacaciones_mes'] for l in liquidaciones_data),
		15: sum(l['bruto_menos_vacaciones'] for l in liquidaciones_data),
		16: sum(l['importe_ss'] for l in liquidaciones_data),
		17: sum(l['total'] for l in liquidaciones_data)
	}
	
	for col_idx, total_value in sum_cols.items():
		cell = ws.cell(row=total_row, column=col_idx, value=total_value)
		cell.font = Font(bold=True)
		cell.border = thin_border
		if col_idx in [9, 10, 11]:
			cell.number_format = number_format
		else:
			cell.number_format = currency_format
			cell.alignment = currency_alignment
	
	ws.cell(row=total_row, column=18).border = thin_border
	
	# Guardar Excel en memoria
	excel_buffer = io.BytesIO()
	wb.save(excel_buffer)
	excel_buffer.seek(0)
	excel_content = excel_buffer.getvalue()
	
	# Nombre del archivo
	filename = f"Liquidaciones_{mes}_{año}_{frappe.utils.now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx"
	
	# Obtener usuarios con rol Asesoría
	asesoria_users = frappe.get_all(
		"Has Role",
		filters={"role": "Asesoría", "parenttype": "User"},
		fields=["parent"]
	)
	
	recipients = [u.parent for u in asesoria_users if u.parent and "@" in u.parent]
	
	if not recipients:
		frappe.throw(_("No se encontraron usuarios con rol Asesoría para enviar el email"))
	
	# Guardar archivo temporal para adjuntar
	file_doc = frappe.get_doc({
		"doctype": "File",
		"file_name": filename,
		"content": excel_content,
		"is_private": 1
	})
	file_doc.insert(ignore_permissions=True)
	
	# Enviar email
	total_importe = sum(l['total'] for l in liquidaciones_data)
	
	frappe.sendmail(
		recipients=recipients,
		subject=f"Liquidaciones de Nóminas - {mes} {año}",
		message=f"""
		<div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
			<h2 style="color: #1F4E79; border-bottom: 2px solid #1F4E79; padding-bottom: 10px;">
				Liquidaciones de Nóminas - {mes} {año}
			</h2>
			
			<p>Se adjunta el archivo Excel con las liquidaciones de nóminas correspondientes a <strong>{mes} {año}</strong>.</p>
			
			<div style="background: #f5f5f5; padding: 15px; border-radius: 5px; margin: 20px 0;">
				<h3 style="margin-top: 0; color: #333;">Resumen:</h3>
				<table style="width: 100%; border-collapse: collapse;">
					<tr>
						<td style="padding: 8px 0; border-bottom: 1px solid #ddd;">Total liquidaciones:</td>
						<td style="padding: 8px 0; border-bottom: 1px solid #ddd; text-align: right; font-weight: bold;">{len(liquidaciones_data)}</td>
					</tr>
					<tr>
						<td style="padding: 8px 0; border-bottom: 1px solid #ddd;">Total horas:</td>
						<td style="padding: 8px 0; border-bottom: 1px solid #ddd; text-align: right; font-weight: bold;">{sum(l['total_horas'] for l in liquidaciones_data):.2f}h</td>
					</tr>
					<tr>
						<td style="padding: 8px 0;">Importe total:</td>
						<td style="padding: 8px 0; text-align: right; font-weight: bold; color: #1F4E79; font-size: 1.2em;">{total_importe:,.2f} €</td>
					</tr>
				</table>
			</div>
			
			<p style="color: #666; font-size: 0.9em;">
				Este email ha sido generado automáticamente desde el Portal RRHH.<br>
				Fecha de envío: {frappe.utils.now_datetime().strftime('%d/%m/%Y %H:%M')}
			</p>
		</div>
		""",
		attachments=[{
			"fname": filename,
			"fcontent": excel_content
		}],
		now=True
	)
	
	# Actualizar estado de las liquidaciones
	liquidaciones_actualizadas = []
	for liq_id in liquidaciones_ids:
		doc = frappe.get_doc(DOCTYPE_NAME, liq_id)
		doc.db_set("estado", "Enviado a Asesoría", update_modified=True)
		doc.db_set("fecha_envio_asesoria", today(), update_modified=False)
		liquidaciones_actualizadas.append(liq_id)
	
	frappe.db.commit()
	
	return {
		"success": True,
		"message": _("Se han enviado {0} liquidaciones a Asesoría").format(len(liquidaciones_actualizadas)),
		"liquidaciones": liquidaciones_actualizadas,
		"recipients": recipients,
		"filename": filename
	}


@frappe.whitelist()
def get_liquidaciones_validadas_mes(mes, año):
	"""
	Obtiene las liquidaciones validadas (docstatus=1) que aún no se han enviado a asesoría.
	"""
	liquidaciones = frappe.get_all(DOCTYPE_NAME,
		filters={
			"mes": mes,
			"año": int(año),
			"docstatus": 1,
			"estado": "Liquidado"
		},
		fields=[
			"name", "employee", "employee_name", "dni_nie", "designation",
			"course", "company", "mes", "año",
			"horas_normales", "horas_extras", "total_horas",
			"bruto", "vacaciones_mes", "bruto_menos_vacaciones",
			"importe_ss", "total", "es_ultimo_mes",
			"source_document", "source_document_type"
		],
		order_by="employee_name asc, course asc"
	)
	
	# Añadir course_display y provincia
	for liq in liquidaciones:
		liq['course_display'] = frappe.db.get_value('Course', liq['course'], 'custom_display_identifier') or liq['course']
		
		# Obtener provincia
		if liq.get('source_document_type') == "Job Offer" and liq.get('source_document'):
			liq['provincia'] = frappe.db.get_value("Job Offer", liq['source_document'], "custom_provincia") or ""
		elif liq.get('source_document_type') == "Modificaciones RRHH" and liq.get('source_document'):
			liq['provincia'] = frappe.db.get_value("Modificaciones RRHH", liq['source_document'], "custom_provincia") or ""
		else:
			liq['provincia'] = ""
	
	return {
		'liquidaciones': liquidaciones,
		'total': len(liquidaciones),
		'total_importe': sum(flt(l['total']) for l in liquidaciones)
	}
